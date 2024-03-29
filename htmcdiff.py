# Copyright (c) 2019-2023 Kirill 'Kolyat' Kiselnikov
# This file is the part of medutils, released under modified MIT license
# See the file LICENSE included in this distribution

import openpyxl
import collections


def find_difference(primary_wb_obj, secondary_wb_name):
    print(f'\n{secondary_wb_name}')
    print('Preparing data...', end=' ', flush=True)
    primary_data = primary_wb_obj['МИС']
    secondary_wb = openpyxl.load_workbook(f'{secondary_wb_name}.xlsx')
    secondary_data = secondary_wb['Лист2']
    print('Done')

    print('Searching for difference...', end=' ', flush=True)
    primary_set = set()
    for row in primary_data.iter_rows(values_only=True):
        primary_set.add(row[4])
    secondary_set = set()
    for row in secondary_data.iter_rows(values_only=True):
        secondary_set.add(row[1])
    set_diff = sorted(secondary_set - primary_set)
    total = len(set_diff)
    print('Done')

    print(f'Writing results to {secondary_wb_name}.xlsx...',
          end=' ', flush=True)
    sheet = secondary_wb.create_sheet(title=secondary_wb_name)
    for i in range(total):
        sheet.cell(column=1, row=i+1, value=set_diff[i])
    secondary_wb.save(f'{secondary_wb_name}.xlsx')
    print('Done')


def count_unique(wb_obj, wb_filename, sheet_name):
    print(f'\nCounting unique operations in {sheet_name}...',
          end=' ', flush=True)
    data = wb_obj[sheet_name]
    counter = collections.Counter()
    for row in data.iter_rows(values_only=True):
        counter.update([row[3]])
    print('Done')

    print(f'Writing results to {wb_filename}...', end=' ', flush=True)
    sheet = wb_obj.create_sheet(title=f'{sheet_name} операции')
    row = 1
    for k, v in counter.items():
        sheet.cell(column=1, row=row, value=k)
        sheet.cell(column=2, row=row, value=v)
        row += 1
    wb_obj.save(wb_filename)
    print('Done')


if __name__ == '__main__':
    MIS_FILENAME = 'МИС.xlsx'
    mis_wb = openpyxl.load_workbook(MIS_FILENAME)
    find_difference(mis_wb, 'ВМП ФЕД')
    find_difference(mis_wb, 'ВМП ОМС')
    count_unique(mis_wb, MIS_FILENAME, 'ВМП')
    count_unique(mis_wb, MIS_FILENAME, 'неВМП')
